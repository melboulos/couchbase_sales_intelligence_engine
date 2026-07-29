# =====================================================
# BUILD AE CALL LIST
# Couchbase Sales Intelligence Engine
#
# DEBUG VERSION — checkpoint prints added after every
# major step to pinpoint exactly where execution stops.
# =====================================================

import sys
import os
import pandas as pd
import ast
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties
from openpyxl.drawing.image import Image as XLImage


def checkpoint(msg):
    print(f">>> CHECKPOINT: {msg}", flush=True)


INPUT_FILE = "output/report1784905185024_Scored_FINAL.xlsx"
OUTPUT_FILE = "output/AE_Call_List.xlsx"
# Real, licensed Couchbase brand banner (768x192, 4:1) - not
# generated or reconstructed here. If this file isn't present, the
# splash banner row is simply left blank rather than failing the
# whole build - see the image-insertion block in the Top 20 sheet
# section for the fallback.
LOGO_FILE = "assets/couchbase_banner.png"


def parse_field(value):
    if isinstance(value, (list, dict)):
        return value
    if not isinstance(value, str) or not value.strip():
        return []
    try:
        return ast.literal_eval(value)
    except (ValueError, SyntaxError):
        return value


def flatten_bullet_list(value):
    parsed = parse_field(value)
    if isinstance(parsed, list):
        if not parsed:
            return "(none noted)"
        return "\n".join(f"\u2022 {item}" for item in parsed)
    if isinstance(parsed, str) and parsed.strip():
        return parsed
    return "(none noted)"


def flatten_discovery_progression(value):
    parsed = parse_field(value)
    if not isinstance(parsed, list) or not parsed:
        return "(none noted)"
    lines = []
    for phase in parsed:
        if not isinstance(phase, dict):
            continue
        phase_label = phase.get("phase", "")
        objective = phase.get("objective", "")
        questions = phase.get("questions", [])
        lines.append(f"{phase_label} \u2014 {objective}")
        for q in questions:
            lines.append(f"    \u2022 {q}")
        lines.append("")
    return "\n".join(lines).strip()


checkpoint("Script started")

print("Loading scored accounts...", flush=True)
accounts = pd.read_excel(INPUT_FILE)
print(f"Loaded {len(accounts)} accounts", flush=True)

checkpoint("Loaded input file")

has_intelligence = accounts["engineering_implications"].notna() & (
    accounts["engineering_implications"].astype(str).str.strip() != ""
)
call_list = accounts[has_intelligence].copy()
call_list = call_list.sort_values("overall_coi", ascending=False).reset_index(drop=True)
print(f"Accounts with validated LLM intelligence: {len(call_list)}", flush=True)

checkpoint("Filtered call_list")

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=16)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Arial", bold=True, size=17, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
KPI_VALUE_FONT = Font(name="Arial", bold=True, size=27, color="1F3864")
LABEL_FONT = Font(name="Arial", bold=True, size=13, color="2F5496")
BODY_FONT = Font(name="Arial", size=13)
LINK_FONT = Font(name="Arial", size=14, color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
BOX_BORDER = Border(
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
)

# Call Briefs-specific fonts, sized larger than the shared constants
# above. Kept separate (not just bumping TITLE_FONT/LABEL_FONT/
# BODY_FONT) because those are also used on Top 20 Accounts and
# Full Landscape - reusing them here would have resized those
# sheets too as an unintended side effect.
BRIEF_TITLE_FONT = Font(name="Arial", bold=True, size=19, color="FFFFFF")
BRIEF_LABEL_FONT = Font(name="Arial", bold=True, size=15, color="2F5496")
BRIEF_BODY_FONT = Font(name="Arial", size=15)
BRIEF_LINK_FONT = Font(name="Arial", size=15, color="0563C1", underline="single")
BRIEF_MARKER_FONT = Font(name="Arial", size=13, italic=True, color="7F7F7F")
BRIEF_CANVAS_HEADER_FONT = Font(name="Arial", bold=True, size=14, color="1F3864")
BRIEF_PRESSURE_BAR_FONT = Font(name="Courier New", size=13)

# Overview-specific fonts, also kept separate from the shared
# constants above for the same reason as the BRIEF_ fonts - HEADER_FONT
# and LINK_FONT are also used on Full Landscape and Top 20 Accounts.
OVERVIEW_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=16)
OVERVIEW_BODY_FONT = Font(name="Arial", size=15)
OVERVIEW_LINK_FONT = Font(name="Arial", size=15, color="0563C1", underline="single")

# SIP/Top 20-specific fonts. BODY_FONT/LINK_FONT above are still used
# by Full Landscape's data rows, which weren't asked to change here -
# bumping those shared constants directly would have resized Full
# Landscape's table body as an unintended side effect.
SIP_BODY_FONT = Font(name="Arial", size=16)
SIP_LINK_FONT = Font(name="Arial", size=16, color="0563C1", underline="single")

checkpoint("Styles defined")

industry_summary = (
    accounts
    .groupby("industry")
    .agg(
        total_accounts=("Account Name", "count"),
        tier_1=("priority_tier", lambda s: (s == "Tier 1 Strategic").sum()),
        tier_2=("priority_tier", lambda s: (s == "Tier 2 Strong Target").sum()),
        tier_3=("priority_tier", lambda s: (s == "Tier 3 Nurture").sum())
    )
    .reset_index()
)

# Averaging COI across ALL accounts (including Tier 4, usually the
# large majority) produced a heavily diluted, confusing number - an
# industry with 950 Tier 4 accounts and 50 strong Tier 1-3 accounts
# would show a low blended average that hides the real opportunity
# quality. Computed separately here, over ONLY the actionable
# (Tier 1-3) accounts per industry - the same population as the call
# list itself - so this reflects "how strong are the accounts we'd
# actually call" rather than a number diluted by accounts nobody was
# ever going to call anyway.
actionable_avg_coi = (
    call_list
    .groupby("industry")["overall_coi"]
    .mean()
    .round(1)
    .reset_index()
    .rename(columns={"overall_coi": "avg_coi"})
)
industry_summary = industry_summary.merge(actionable_avg_coi, on="industry", how="left")
industry_summary["avg_coi"] = industry_summary["avg_coi"].fillna(0)

industry_summary["actionable_pct"] = (
    100 * (industry_summary["tier_1"] + industry_summary["tier_2"] + industry_summary["tier_3"])
    / industry_summary["total_accounts"]
).round(1)

# merge() appends the new column at the end - explicitly restore its
# intended position (right after total_accounts, before the tier
# columns), since the conditional color formatting below is
# hardcoded to column C and would silently format the wrong column
# otherwise.
industry_summary = industry_summary[
    ["industry", "total_accounts", "avg_coi", "actionable_pct", "tier_1", "tier_2", "tier_3"]
]

checkpoint("industry_summary computed")

industry_summary = industry_summary.sort_values("tier_1", ascending=False).reset_index(drop=True)
industry_summary = industry_summary.head(15)
industry_summary = industry_summary.rename(columns={
    "industry": "Industry",
    "total_accounts": "Total Accounts",
    "avg_coi": "Avg COI (Actionable)",
    "actionable_pct": "Actionable %",
    "tier_1": "Tier 1",
    "tier_2": "Tier 2",
    "tier_3": "Tier 3"
})

checkpoint("industry_summary formatted, about to write initial xlsx")

industry_summary.to_excel(OUTPUT_FILE, index=False, sheet_name="Full Landscape", startrow=6)

checkpoint("Initial pandas to_excel write complete")

wb = load_workbook(OUTPUT_FILE)
checkpoint("Reloaded workbook with openpyxl")

ws_summary = wb["Full Landscape"]
checkpoint("Got Summary worksheet handle")

ws_summary.merge_cells("A1:J1")
title_cell = ws_summary["A1"]
title_cell.value = "Couchbase Sales Intelligence \u2014 Account Landscape Summary"
title_cell.font = TITLE_FONT
title_cell.fill = TITLE_FILL
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_summary.row_dimensions[1].height = 32

checkpoint("Title bar written")

total_scored = len(accounts)
total_qualified = len(call_list)
total_tier1 = int((accounts["priority_tier"] == "Tier 1 Strategic").sum())
total_tier2 = int((accounts["priority_tier"] == "Tier 2 Strong Target").sum())
actionable_pct_fraction = (total_qualified / total_scored) if total_scored > 0 else 0

kpis = [
    ("Total Accounts Scored", total_scored),
    ("Actionable Accounts (Tier 1\u20133)", total_qualified),
    ("Actionable %", actionable_pct_fraction),
    ("Tier 1 Strategic", total_tier1),
    ("Tier 2 Strong Target", total_tier2)
]

kpi_col = 1
for label, value in kpis:
    label_cell = ws_summary.cell(row=3, column=kpi_col)
    value_cell = ws_summary.cell(row=4, column=kpi_col)
    label_cell.value = label
    label_cell.font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    label_cell.fill = HEADER_FILL
    label_cell.alignment = CENTER
    value_cell.value = value
    value_cell.font = KPI_VALUE_FONT
    value_cell.alignment = CENTER
    if label == "Actionable %":
        # Real number (a fraction), formatted with Excel's native
        # percent format - not a Python string with a literal "%"
        # appended, which is exactly what triggers Excel's genuine
        # "Number Stored as Text" warning (confirmed via screenshot).
        value_cell.number_format = "0.0%"
    ws_summary.merge_cells(start_row=3, start_column=kpi_col, end_row=3, end_column=kpi_col + 1)
    ws_summary.merge_cells(start_row=4, start_column=kpi_col, end_row=4, end_column=kpi_col + 1)
    ws_summary.row_dimensions[3].height = 26
    ws_summary.row_dimensions[4].height = 34
    kpi_col += 2

checkpoint("KPI row written")

header_row = 7
n_cols = len(industry_summary.columns)
# Table (header + data) spans only its real A:G columns - the page
# title and KPI row above are wider (A:J) but that's fine on its own;
# it was specifically the table's own header extending past its real
# data that looked wrong, since a table header implies a table below
# it, not a page-level summary bar.
for col_idx in range(1, n_cols + 1):
    cell = ws_summary.cell(row=header_row, column=col_idx)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BOX_BORDER
ws_summary.row_dimensions[header_row].height = 28

summary_widths = {"A": 32, "B": 21, "C": 30, "D": 19, "E": 12, "F": 12, "G": 12, "H": 14, "I": 14, "J": 13}
for col, width in summary_widths.items():
    ws_summary.column_dimensions[col].width = width

first_data_row = header_row + 1
last_data_row = header_row + len(industry_summary)

for row_idx in range(first_data_row, last_data_row + 1):
    ws_summary.row_dimensions[row_idx].height = 24
    for col_idx in range(1, n_cols + 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx)
        cell.border = BOX_BORDER
        cell.font = BODY_FONT
        if col_idx > 1:
            cell.alignment = CENTER

checkpoint("Summary table styled")

if last_data_row > first_data_row:
    avg_coi_range = f"C{first_data_row}:C{last_data_row}"
    ws_summary.conditional_formatting.add(
        avg_coi_range,
        ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            mid_type="percentile", mid_value=50, mid_color="FFF3B0",
            end_type="max", end_color="F2A9A5"
        )
    )
    for col_letter in ["D", "E", "F", "G"]:
        cell_range = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
        ws_summary.conditional_formatting.add(
            cell_range,
            ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                mid_type="percentile", mid_value=50, mid_color="FFF3B0",
                end_type="max", end_color="F2A9A5"
            )
        )

checkpoint("Conditional formatting applied")

chart = BarChart()
chart.type = "bar"
chart.title = "Tier 1 Strategic Accounts by Industry"
chart.style = 10
chart.y_axis.title = "Industry"
chart.x_axis.title = "Tier 1 Account Count"
chart.height = 10
chart.width = 20

data = Reference(ws_summary, min_col=5, max_col=5, min_row=header_row, max_row=last_data_row)
categories = Reference(ws_summary, min_col=1, max_col=1, min_row=first_data_row, max_row=last_data_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
chart.legend = None

chart_anchor_row = last_data_row + 3
ws_summary.add_chart(chart, f"A{chart_anchor_row}")

checkpoint("Chart added")

ws_summary.freeze_panes = f"A{first_data_row}"

checkpoint("About to create Overview sheet")

ws_overview = wb.create_sheet("Overview")

overview_headers = ["Account Name", "COI Score", "Priority Tier", "Industry", "Business Model", "Account Owner"]
for col_idx, header in enumerate(overview_headers, start=1):
    cell = ws_overview.cell(row=1, column=col_idx)
    cell.value = header
    cell.font = OVERVIEW_HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(vertical="center")
ws_overview.row_dimensions[1].height = 27

# Lets reps filter/sort by owner, tier, industry, or COI directly in
# Excel instead of scrolling - dropdown arrows appear on every header.
ws_overview.auto_filter.ref = f"A1:{get_column_letter(len(overview_headers))}1"

# Account Name widened from 28 to 34 based on real distribution (95th
# percentile is 36 chars, median 17 - a handful of ~90-char government/
# university names are genuine long-tail outliers, not the norm; those
# wrap to a second line below rather than forcing the column absurdly
# wide for everyone else). Other columns widened proportionally.
overview_widths = {"A": 34, "B": 15, "C": 25, "D": 27, "E": 27, "F": 23}
for col, width in overview_widths.items():
    ws_overview.column_dimensions[col].width = width

# Account Name row height is computed from real name length, same
# principle used everywhere else in this script - most names fit on
# one line at this width, the rare long outlier wraps instead of
# truncating or forcing every row taller than it needs to be.
CHARS_PER_LINE_AT_OVERVIEW_A_WIDTH = 30

# NOTE: a "Confidence" column (Web-Verified / Not Company-Verified) was
# tried here on 2026-07-29 and reverted the same day. Real data showed
# 99.7% of accounts (3,567 of 3,579) are Web-Verified - since the
# pipeline attempts web grounding for nearly every qualifying account
# by design, that flag isn't a decision-relevant signal for a rep, just
# confirmation the pipeline did its job. Not worth a column for that.
# If this is revisited, the useful version would only surface the rare
# Not Company-Verified exception (11 accounts), not all three states -
# see caveat_marker/web_search_marker in the Call Briefs section below
# for the equivalent logic on that sheet.

# Priority tier -> fill color, for at-a-glance scanning without reading text
TIER_FILLS = {
    "Tier 1 Strategic": PatternFill(start_color="C0DD97", end_color="C0DD97", fill_type="solid"),
    "Tier 2 Strong Target": PatternFill(start_color="B5D4F4", end_color="B5D4F4", fill_type="solid"),
    "Tier 3 Nurture": PatternFill(start_color="FAC775", end_color="FAC775", fill_type="solid"),
    "Tier 4 Monitor": PatternFill(start_color="D3D1C7", end_color="D3D1C7", fill_type="solid"),
}

account_brief_rows = []

for i, row in call_list.iterrows():
    excel_row = i + 2
    ws_overview.row_dimensions[excel_row].height = 23
    ws_overview.cell(row=excel_row, column=2, value=row.get("overall_coi", "")).font = OVERVIEW_BODY_FONT
    tier_cell = ws_overview.cell(row=excel_row, column=3, value=row.get("priority_tier", ""))
    tier_cell.font = OVERVIEW_BODY_FONT
    tier_fill = TIER_FILLS.get(row.get("priority_tier", ""))
    if tier_fill:
        tier_cell.fill = tier_fill
    ws_overview.cell(row=excel_row, column=4, value=row.get("industry", "")).font = OVERVIEW_BODY_FONT
    ws_overview.cell(row=excel_row, column=5, value=row.get("business_model", "")).font = OVERVIEW_BODY_FONT
    ws_overview.cell(row=excel_row, column=6, value=row.get("Account Owner", "")).font = OVERVIEW_BODY_FONT
    for col_idx in range(2, 7):
        cell = ws_overview.cell(row=excel_row, column=col_idx)
        cell.alignment = CENTER
        cell.border = BOX_BORDER

ws_overview.freeze_panes = "A2"

checkpoint(f"Overview sheet built with {len(call_list)} rows")

ws_briefs = wb.create_sheet("Call Briefs")
ws_briefs.column_dimensions["A"].width = 30
ws_briefs.column_dimensions["B"].width = 105

def make_bar(value, max_value, width=10):
    """
    Renders a unicode block bar honestly proportional to a REAL
    score value out of its real, documented max (see
    modules/scoring_engine.py for the actual caps: workload=40,
    database=30, realtime=15, technical=10, company=5). No
    fabricated dimensions - only components that actually exist in
    the deterministic scoring engine.
    """
    try:
        value = float(value)
    except (TypeError, ValueError):
        value = 0
    if max_value <= 0:
        filled = 0
    else:
        filled = round((value / max_value) * width)
    filled = max(0, min(width, filled))
    return "\u2588" * filled + "\u2591" * (width - filled)


def build_opportunity_pressure_bars(row):
    components = [
        ("Workload Fit", row.get("workload_fit_points", 0), 40),
        ("Database Opportunity", row.get("database_opportunity_points", 0), 30),
        ("Real-Time Need", row.get("realtime_points", 0), 15),
        ("Technical Environment", row.get("technical_environment_points", 0), 10),
        ("Company Scale", row.get("company_context_points", 0), 5),
    ]
    lines = [f"{name:<22} {make_bar(value, max_val)}" for name, value, max_val in components]
    return "\n".join(lines)


def build_discovery_checklist(row):
    parsed = parse_field(row.get("discovery_progression", ""))
    if not isinstance(parsed, list):
        return "(none noted)"
    lines = []
    for phase in parsed:
        if isinstance(phase, dict):
            obj = phase.get("objective", "")
            if obj:
                lines.append(f"\u2610 {obj}")
    return "\n".join(lines) if lines else "(none noted)"


def build_research_confidence(row):
    if row.get("llm_narrative_caveated") in (True, 1, 1.0):
        return "\u26A0 Not company-verified - built on an assigned category, not confirmed company knowledge."
    if row.get("llm_used_web_search") in (True, 1, 1.0):
        return "\U0001F50D Web-verified - grounded in a real, live search result."
    return "Verified from model recognition (no live search performed for this account)."


CANVAS_FIELD_ORDER = [
    ("\U0001F3AF Why This Account", lambda r: (flatten_bullet_list(r.get("engineering_implications", "")).split("\n")[0]
                                                 if r.get("engineering_implications") else "(no engineering implications noted)")),
    ("\U0001F4BC Business Context", lambda r: (
        f"Industry: {r.get('industry','Unknown')}  |  "
        f"Business Model: {r.get('business_model','Unknown')}  |  "
        f"Priority: {r.get('priority_tier','Unknown')}"
    )),
    ("\u2699\ufe0f Likely Workload", lambda r: str(r.get("workload_profile", "Unknown"))),
    ("\U0001F525 Engineering Pressures", build_opportunity_pressure_bars),
    ("\u2753 Discovery Objectives", build_discovery_checklist),
    ("\u2705 Research Confidence", build_research_confidence),
]


FIELD_ORDER = [
    ("Account Owner", lambda r: r.get("Account Owner", "")),
    ("Industry", lambda r: r.get("industry", "")),
    ("Business Model", lambda r: r.get("business_model", "")),
    ("Workload Profile", lambda r: r.get("workload_profile", "")),
    ("\u2699\ufe0f Engineering Implications", lambda r: flatten_bullet_list(r.get("engineering_implications", ""))),
    ("\U0001F4A1 Couchbase Point of View", lambda r: r.get("couchbase_point_of_view", "") or "(none noted)"),
    ("\U0001F6A9 Technical Risks to Validate", lambda r: flatten_bullet_list(r.get("technical_risks_to_validate", ""))),
    ("\u2753 Discovery Questions", lambda r: flatten_discovery_progression(r.get("discovery_progression", ""))),
    ("Missing Information", lambda r: flatten_bullet_list(r.get("missing_information", "")))
]

current_row = 1
checkpoint("Starting Call Briefs loop (this is the biggest step — 320 accounts)")

for i, row in call_list.iterrows():
    if i % 50 == 0:
        checkpoint(f"Call Briefs progress: {i}/{len(call_list)}")

    title_row = current_row
    account_brief_rows.append((row.get("Account Name", ""), title_row))

    ws_briefs.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=2)
    title_cell = ws_briefs.cell(row=title_row, column=1)
    caveat_marker = (
        "\u26A0 NOT COMPANY-VERIFIED"
        if row.get("llm_narrative_caveated") in (True, 1, 1.0)
        else ""
    )
    web_search_marker = (
        "\U0001F50D Web-Verified"
        if row.get("llm_used_web_search") in (True, 1, 1.0)
        else ""
    )
    llm_score_display = row.get("llm_total_score", "")
    llm_score_default_marker = (
        " (default)"
        if row.get("llm_score_is_default") in (True, 1, 1.0)
        else ""
    )
    title_cell.value = (
        f"{row.get('Account Name', '')}   "
        f"\u2014  COI {row.get('overall_coi', '')}  "
        f"\u2014  LLM Score {llm_score_display}{llm_score_default_marker}  "
        f"\u2014  {row.get('priority_tier', '')}"
    )
    title_cell.font = BRIEF_TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(vertical="center", indent=1)
    ws_briefs.row_dimensions[title_row].height = 36

    current_row += 1

    # Verification markers get their own row - moved out of the title
    # after finding the combined title could overflow its cell width
    # (confirmed via real font-metrics check: a title with the LLM
    # score plus even one marker exceeded the available cell width
    # in the common case, not just a rare worst-case combination).
    markers = "   ".join(m for m in [caveat_marker, web_search_marker] if m)
    if markers:
        marker_cell = ws_briefs.cell(row=current_row, column=1)
        ws_briefs.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        marker_cell.value = markers
        marker_cell.font = BRIEF_MARKER_FONT
        marker_cell.alignment = Alignment(vertical="center", indent=1)
        ws_briefs.row_dimensions[current_row].height = 22
        current_row += 1

    back_cell = ws_briefs.cell(row=current_row, column=1)
    back_cell.value = "\u2191 Back to Overview"
    back_cell.hyperlink = "#Overview!A1"
    back_cell.font = BRIEF_LINK_FONT
    back_cell.alignment = Alignment(vertical="center", indent=1)
    ws_briefs.row_dimensions[current_row].height = 20

    current_row += 1

    canvas_header_cell = ws_briefs.cell(row=current_row, column=1)
    ws_briefs.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    canvas_header_cell.value = "\U0001F4CB TECHNICAL OPPORTUNITY CANVAS"
    canvas_header_cell.font = BRIEF_CANVAS_HEADER_FONT
    canvas_header_cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    canvas_header_cell.alignment = Alignment(vertical="center", indent=1)
    ws_briefs.row_dimensions[current_row].height = 25
    current_row += 1

    for label, getter in CANVAS_FIELD_ORDER:
        label_cell = ws_briefs.cell(row=current_row, column=1)
        value_cell = ws_briefs.cell(row=current_row, column=2)

        label_cell.value = label
        label_cell.font = BRIEF_LABEL_FONT
        label_cell.fill = PatternFill(start_color="F2F6FB", end_color="F2F6FB", fill_type="solid")
        label_cell.alignment = Alignment(vertical="top", wrap_text=True)
        label_cell.border = THIN_BORDER

        value_cell.value = getter(row)
        value_cell.font = BRIEF_PRESSURE_BAR_FONT if label.startswith("\U0001F525") else BRIEF_BODY_FONT
        value_cell.fill = PatternFill(start_color="F2F6FB", end_color="F2F6FB", fill_type="solid")
        value_cell.alignment = WRAP
        value_cell.border = THIN_BORDER

        text_len = len(str(value_cell.value))
        line_estimate = max(1, text_len // 62 + str(value_cell.value).count("\n") + 1)
        ws_briefs.row_dimensions[current_row].height = min(max(24, line_estimate * 21), 360)

        current_row += 1

    current_row += 1

    for label, getter in FIELD_ORDER:
        label_cell = ws_briefs.cell(row=current_row, column=1)
        value_cell = ws_briefs.cell(row=current_row, column=2)

        label_cell.value = label
        label_cell.font = BRIEF_LABEL_FONT
        label_cell.alignment = Alignment(vertical="top", wrap_text=True)
        label_cell.border = THIN_BORDER

        value_cell.value = getter(row)
        value_cell.font = BRIEF_BODY_FONT
        value_cell.alignment = WRAP
        value_cell.border = THIN_BORDER

        text_len = len(str(value_cell.value))
        line_estimate = max(1, text_len // 62 + str(value_cell.value).count("\n") + 1)
        ws_briefs.row_dimensions[current_row].height = min(max(24, line_estimate * 21), 360)

        current_row += 1

    current_row += 1

checkpoint("Call Briefs loop complete")

ws_briefs.freeze_panes = "A2"

checkpoint("About to write hyperlinks")

for i in range(len(call_list)):
    row_num = i + 2
    account_name, brief_row = account_brief_rows[i]
    name_cell = ws_overview.cell(row=row_num, column=1)
    name_cell.value = account_name
    name_cell.hyperlink = f"#'Call Briefs'!A{brief_row}"
    name_cell.font = OVERVIEW_LINK_FONT
    name_cell.alignment = Alignment(wrap_text=True, vertical="center")
    name_cell.border = BOX_BORDER

    line_estimate = max(1, len(str(account_name)) // CHARS_PER_LINE_AT_OVERVIEW_A_WIDTH + 1)
    needed_height = min(line_estimate * 21, 70)
    if needed_height > ws_overview.row_dimensions[row_num].height:
        ws_overview.row_dimensions[row_num].height = needed_height

checkpoint("Hyperlinks written")

checkpoint("Building Top 20 Accounts sheet")

account_brief_row_lookup = dict(account_brief_rows)

# Deterministic, not LLM-generated - a reasonable starting suggestion
# for who to reach out to first, based purely on the workload
# category. Explicitly a suggestion, not a claim about the specific
# account - the rep's own judgment about the actual account always
# wins over a generic per-category default.
PERSONA_BY_WORKLOAD_PROFILE = {
    "payment_platform": "Director of Payments / Head of Fraud Engineering",
    "api_platform": "VP of Engineering / Head of Platform",
    "saas_platform": "VP of Engineering / CTO",
    "customer_application": "VP of Product Engineering",
    "mobile_application": "Head of Mobile Engineering",
    "utilities_platform": "Director of IT Operations",
    "media_platform": "VP of AdTech Engineering",
    "logistics_platform": "VP of Supply Chain Technology",
    "retail_platform": "Director of Retail Technology",
    "insurance_platform": "Director of Enterprise Architecture",
    "pharma_device_platform": "Director of IT / Regulatory Systems",
    "telecom_platform": "VP of Network Engineering",
    "media_entertainment_platform": "VP of Streaming Engineering",
}


def get_recommended_contact(workload_profile):
    return PERSONA_BY_WORKLOAD_PROFILE.get(workload_profile, "VP of Engineering")


def get_why_couchbase(pov_text):
    """
    Returns the full couchbase_point_of_view text, unmodified.

    Previously this cut the text down to just the first sentence,
    then capped that sentence at 15 words with "..." - a truncation
    that was tried and explicitly rejected. Row height for this
    column (see CHARS_PER_LINE_AT_E_WIDTH below) is already computed
    dynamically from real text length, so no truncation is needed
    here for the cell to display correctly.
    """
    if not isinstance(pov_text, str) or not pov_text.strip():
        return "(none noted)"
    return pov_text.strip()


def set_chart_fonts(chart, title_size=1600, axis_size=1200):
    """
    openpyxl charts don't inherit worksheet fonts - title/axis text
    needs its own explicit rich-text font size, or it renders at a
    tiny default regardless of chart.height/width. Sizes are in
    hundredths of a point (1600 = 16pt).

    PieChart has no x_axis/y_axis at all (categorical, not axis-
    based) - only bar-style charts do. Checking hasattr before
    accessing them, rather than assuming every chart type has axes.
    """
    if chart.title and chart.title.tx and chart.title.tx.rich:
        for para in chart.title.tx.rich.p:
            if para.pPr is None:
                para.pPr = ParagraphProperties()
            para.pPr.defRPr = CharacterProperties(sz=title_size, b=True)
    for axis_name in ["x_axis", "y_axis"]:
        axis = getattr(chart, axis_name, None)
        if axis is not None:
            axis.txPr = RichText(
                p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=axis_size)))]
            )
    if hasattr(chart, "legend") and chart.legend is not None:
        chart.legend.txPr = RichText(
            p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=axis_size)))]
        )


top_20 = call_list.sort_values("overall_coi", ascending=False).head(20).copy()
top_20["why_couchbase"] = top_20["couchbase_point_of_view"].apply(get_why_couchbase)
top_20["recommended_contact"] = top_20["workload_profile"].apply(get_recommended_contact)

driver_counts = call_list["workload_profile"].value_counts()
driver_pcts = (driver_counts / len(call_list) * 100).round(1).head(6)

tier_counts = accounts["priority_tier"].value_counts()
tier_order = ["Tier 1 Strategic", "Tier 2 Strong Target", "Tier 3 Nurture", "Tier 4 Monitor"]
tier_counts = tier_counts.reindex(tier_order).fillna(0).astype(int)

industry_opportunity = call_list.groupby("industry").size().sort_values(ascending=False).head(10)

ws_top20 = wb.create_sheet("Sales Intelligence Platform", 0)
ws_top20.sheet_view.showGridLines = False

ws_top20.column_dimensions["A"].width = 11
ws_top20.column_dimensions["B"].width = 32
ws_top20.column_dimensions["C"].width = 10
ws_top20.column_dimensions["D"].width = 24
ws_top20.column_dimensions["E"].width = 120
ws_top20.column_dimensions["F"].width = 34

# Splash banner, three tiers:
#   Row 1 - real Couchbase brand banner image (assets/couchbase_banner.png,
#           4:1 aspect, sized to preserve that ratio rather than stretched
#           to fill the full table width and getting distorted/blurry).
#   Row 2 - project title
#   Row 3 - sheet subtitle
# table_header_row is the only thing that needed to change to make
# room - everything below it (table_end_row, drivers_header_row,
# chart anchors) is computed relative to it, not hardcoded.
if os.path.exists(LOGO_FILE):
    banner_img = XLImage(LOGO_FILE)
    banner_img.width = 600
    banner_img.height = 150  # preserves the source's real 4:1 ratio
    ws_top20.add_image(banner_img, "A1")
    checkpoint(f"Inserted real banner image from {LOGO_FILE}")
else:
    checkpoint(f"WARNING: {LOGO_FILE} not found - splash row left blank, not failing the build")
ws_top20.row_dimensions[1].height = 114

ws_top20.merge_cells("A2:F2")
top20_title = ws_top20["A2"]
top20_title.value = "Sales Intelligence Platform"
top20_title.font = Font(name="Arial", bold=True, size=22, color="FFFFFF")
top20_title.fill = TITLE_FILL
top20_title.alignment = Alignment(horizontal="center", vertical="center")
ws_top20.row_dimensions[2].height = 34

ws_top20.merge_cells("A3:F3")
top20_subtitle = ws_top20["A3"]
top20_subtitle.value = "\U0001F3AF Top 20 Accounts to Call"
top20_subtitle.font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
top20_subtitle.fill = HEADER_FILL
top20_subtitle.alignment = Alignment(horizontal="center", vertical="center")
ws_top20.row_dimensions[3].height = 26

table_header_row = 4
headers = ["Rank", "Account", "COI", "Workload Profile", "Why Couchbase", "Recommended First Contact"]
for col_idx, header in enumerate(headers, start=1):
    cell = ws_top20.cell(row=table_header_row, column=col_idx)
    cell.value = header
    cell.font = HEADER_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
    cell.border = BOX_BORDER
ws_top20.row_dimensions[table_header_row].height = 40

# E column is 95 units wide - roughly 90 characters per line at this
# font size/width. Row height is computed from the REAL text length,
# not a fixed guess - a fixed 32pt row was clipping longer sentences
# even though the column itself was wide enough, since wrap_text
# still needs enough row height to actually show every wrapped line.
CHARS_PER_LINE_AT_E_WIDTH = 90

for i, (_, row) in enumerate(top_20.iterrows()):
    r = table_header_row + 1 + i
    rank_cell = ws_top20.cell(row=r, column=1, value=i + 1)
    rank_cell.font = SIP_BODY_FONT
    rank_cell.alignment = Alignment(horizontal="center", vertical="top")
    rank_cell.border = BOX_BORDER

    account_cell = ws_top20.cell(row=r, column=2, value=row.get("Account Name", ""))
    account_cell.font = SIP_LINK_FONT
    account_cell.alignment = Alignment(vertical="top")
    account_cell.border = BOX_BORDER
    account_name_lookup = row.get("Account Name", "")
    if account_name_lookup in account_brief_row_lookup:
        account_cell.hyperlink = f"#'Call Briefs'!A{account_brief_row_lookup[account_name_lookup]}"

    coi_cell = ws_top20.cell(row=r, column=3, value=row.get("overall_coi", ""))
    coi_cell.font = SIP_BODY_FONT
    coi_cell.alignment = Alignment(horizontal="left", vertical="top")
    coi_cell.border = BOX_BORDER

    workload_cell = ws_top20.cell(row=r, column=4, value=row.get("workload_profile", ""))
    workload_cell.font = SIP_BODY_FONT
    workload_cell.alignment = Alignment(vertical="top")
    workload_cell.border = BOX_BORDER

    why_text = row.get("why_couchbase", "")
    why_cell = ws_top20.cell(row=r, column=5, value=why_text)
    why_cell.font = SIP_BODY_FONT
    why_cell.alignment = WRAP
    why_cell.border = BOX_BORDER

    contact_cell = ws_top20.cell(row=r, column=6, value=row.get("recommended_contact", ""))
    contact_cell.font = SIP_BODY_FONT
    contact_cell.alignment = WRAP
    contact_cell.border = BOX_BORDER

    line_estimate = max(1, (len(str(why_text)) // CHARS_PER_LINE_AT_E_WIDTH) + 1)
    ws_top20.row_dimensions[r].height = max(28, line_estimate * 25)

table_end_row = table_header_row + len(top_20)

coi_range = f"C{table_header_row + 1}:C{table_end_row}"
ws_top20.conditional_formatting.add(
    coi_range,
    ColorScaleRule(
        start_type="min", start_color="FFFFFF",
        mid_type="percentile", mid_value=50, mid_color="FFF3B0",
        end_type="max", end_color="F2A9A5"
    )
)

drivers_header_row = table_end_row + 3
ws_top20.merge_cells(start_row=drivers_header_row, start_column=1, end_row=drivers_header_row, end_column=2)
drivers_title = ws_top20.cell(row=drivers_header_row, column=1)
drivers_title.value = "\U0001F4CA Top Opportunity Drivers"
drivers_title.font = Font(name="Arial", bold=True, size=16, color="1F3864")

for i, (profile, pct) in enumerate(driver_pcts.items()):
    r = drivers_header_row + 1 + i
    ws_top20.cell(row=r, column=1, value=f"\u2022 {pct}%").font = Font(name="Arial", bold=True, size=16)
    ws_top20.cell(row=r, column=2, value=str(profile)).font = SIP_BODY_FONT

# Tier Distribution chart sits BELOW the main table, in the empty
# whitespace at columns C onward next to the Top Opportunity Drivers
# list (that list only occupies columns A-B, so no collision).
tier_data_row = drivers_header_row
tier_data_col = 3
ws_top20.cell(row=tier_data_row, column=tier_data_col, value="Tier")
ws_top20.cell(row=tier_data_row, column=tier_data_col + 1, value="Count")
for i, (tier, count) in enumerate(tier_counts.items()):
    ws_top20.cell(row=tier_data_row + 1 + i, column=tier_data_col, value=tier)
    ws_top20.cell(row=tier_data_row + 1 + i, column=tier_data_col + 1, value=int(count))

tier_chart = PieChart()
tier_chart.title = "Tier Distribution"
tier_chart_data = Reference(ws_top20, min_col=tier_data_col + 1, min_row=tier_data_row, max_row=tier_data_row + len(tier_counts))
tier_chart_cats = Reference(ws_top20, min_col=tier_data_col, min_row=tier_data_row + 1, max_row=tier_data_row + len(tier_counts))
tier_chart.add_data(tier_chart_data, titles_from_data=True)
tier_chart.set_categories(tier_chart_cats)
tier_chart.height = 12
tier_chart.width = 20
set_chart_fonts(tier_chart)
ws_top20.add_chart(tier_chart, f"C{tier_data_row}")

charts_start_row = drivers_header_row + len(driver_pcts) + 3

industry_data_row = charts_start_row
ws_top20.cell(row=industry_data_row, column=1, value="Industry")
ws_top20.cell(row=industry_data_row, column=2, value="Actionable Accounts")
for i, (industry, count) in enumerate(industry_opportunity.items()):
    ws_top20.cell(row=industry_data_row + 1 + i, column=1, value=industry)
    ws_top20.cell(row=industry_data_row + 1 + i, column=2, value=int(count))

industry_chart = BarChart()
industry_chart.type = "bar"
industry_chart.title = "Industry Opportunity Distribution"
industry_chart_data = Reference(ws_top20, min_col=2, min_row=industry_data_row, max_row=industry_data_row + len(industry_opportunity))
industry_chart_cats = Reference(ws_top20, min_col=1, min_row=industry_data_row + 1, max_row=industry_data_row + len(industry_opportunity))
industry_chart.add_data(industry_chart_data, titles_from_data=True)
industry_chart.set_categories(industry_chart_cats)
industry_chart.height = 13
industry_chart.width = 24
set_chart_fonts(industry_chart)
# Anchored below the pie chart (same column C, real gap beneath it)
# rather than on top of its own "Industry / Actionable Accounts"
# source table in columns A-B - that same-row/same-column anchor
# was the actual cause of the two overlapping on screen.
bar_chart_anchor_row = tier_data_row + 25
ws_top20.add_chart(industry_chart, f"C{bar_chart_anchor_row}")

checkpoint("Top 20 Accounts sheet complete")

wb.active = wb.sheetnames.index("Sales Intelligence Platform")

checkpoint("About to save workbook — this may take a moment")

wb.save(OUTPUT_FILE)

checkpoint("Workbook saved successfully")

print(f"Saved: {OUTPUT_FILE}", flush=True)
print(f"Summary: {len(industry_summary)} industries, {total_scored} accounts scored", flush=True)
print(f"Overview + Call Briefs: {len(call_list)} qualified accounts", flush=True)

checkpoint("Script finished")
